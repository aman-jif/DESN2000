# Import relevant modules and classes
import datetime, random
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from helper import decimal_trunc

# Create random float values to pass into the sensors
def generate_random_sensor_data():
    sensor_list = ["Temperature", "pH", "Condensation", "Nitrate"]
    for sensor_index, sensor_value in enumerate(sensor_list):
    # Random values
        sensor_value = float(decimal_trunc(random.uniform(20,60), 2))
        sensor_list[sensor_index] = (sensor_value)
    return sensor_list
    
    
# Function to append data from the sensors into the next new row on the spreadsheet    
def add_random_data():
    rand_sensor_values = generate_random_sensor_data()
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    new_data = [timestamp] + rand_sensor_values
    ws.append(new_data)
    wb.save('water_quality_data.xlsx')

# File validity check
file_validity = False
try:
    # Test if the intended excel sheet exists
    wb_check = load_workbook(filename='water_quality_data.xlsx')
    file_validity = True
    print("File found")
except FileNotFoundError:
    # First exception raised for FileNotFound specifically
    print("File not found - Check the name?")
    file_validity = False
except Exception:
    # All other exceptions are handled here
    print("File not found")
    file_validity = False

if file_validity == False:
    # Create new workbook for data
    wb = Workbook()
    ws = wb.active
    ws.title = "Water Quality Data"

    # If new workbook is created, default initialisation of headers
    headers = ['Timestamp', 'Temperature (Sensor 1)', 'pH (Sensor 2)', 'Conductivity (Sensor 3)',
               'Nitrates (Sensor 4)']
    ws.append(headers)

    # Set header column width
    for i, header in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = len(header)
    # Give timestamp column extra width
    ws.column_dimensions[get_column_letter(1)].width = 20
    
    # Stylise all column names to have centered text
    for col in wb.active.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
    
    # Dummy values to test functionality, with the "timestamp" field real-time testing is enabled
    add_random_data()

else:
    # Case where the workbook was found
    wb = load_workbook(filename = "water_quality_data.xlsx")
    ws = wb.active
    # Dummy values to test functionality, with the "timestamp" field real-time testing is enabled
    add_random_data()